# Emit the generated `estMatrix.cons` block for a ticker, from the UNION of
#   (a) BBG_CONSENSUS.txt   — the exported ARCHIVE (never loses a snapshot)
#   (b) Consensus_Portal.xlsm, the LIVE sheet (can overwrite its last row)
# deduped by (ticker, data_as_of). Only FORWARD horizons are consensus.
#
# ── SCHEMA v2 (Aug 2026) ──────────────────────────────────────────────────────
# The workbook is now SELF-DESCRIBING. Each row declares its own slots:
#     metric1..50  / code1..50  / segment1..50  / unit1..50  / scale1..50
#     metric_kpi1..50 / code_kpi1..50 / segment_kpi1..50 / unit_kpi1..50 / scale_kpi1..50
# and carries the values in `<slot>_<horizon>` columns, e.g. `metric12_fq+1`, `kpi3_fy+2`,
# over horizons fq-4..fq+4 and fy-2..fy+3. The period each horizon points at is in the
# `fq+1` / `fy0` / … columns, labelled `2026 Q3 (Fwd)` / `2025 A (Rep)`.
#
# So a metric is addressed by its BLOOMBERG CODE (+ segment id), NOT by slot position —
# slot numbering is per ticker and can drift between exports. That is the whole point of
# this rewrite: the config below is portable to any ticker without reading a header by eye.
#
# The previous version addressed columns by fixed names (`rev_fq+1`, `ebitda_fq0`, `kpi1_…`).
# Those columns no longer exist except for the `kpiN` slots, which is why a stale config
# silently yields an almost-empty matrix. Legacy slot-name configs still work (see resolve()),
# but new tickers should use the code form.
import openpyxl, csv, json, sys, re

import os
# The two Bloomberg sources live in the team's Google Drive, whose drive letter differs per
# machine. Override with the SUMMIT_DOCS environment variable:
#     SUMMIT_DOCS="D:/My Drive/Summit/Docs/0" py emit_matrix.py AMZN map_amzn.json
DOCS = os.environ.get('SUMMIT_DOCS', 'G:/My Drive/Summit/Docs/0')
HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, 'out'); os.makedirs(OUT, exist_ok=True)

TXT = os.path.join(DOCS, 'BBG_CONSENSUS.txt')
WB  = os.path.join(DOCS, 'Consensus_Portal.xlsm')
# The live sheet was renamed BBG_CONSENSUS -> CONSENSUS in the Aug-2026 rebuild; take either.
WB_SHEETS = ['CONSENSUS', 'BBG_CONSENSUS']
TK  = (sys.argv[1] if len(sys.argv) > 1 else "UBER").upper()
MAP = json.load(open(sys.argv[2], encoding='utf-8'))
VALID = json.load(open(sys.argv[3], encoding='utf-8')) if len(sys.argv) > 3 else {}
SEGS = MAP.get('segments', {})     # friendly alias -> Bloomberg segment id

QH = ['fq-4','fq-3','fq-2','fq-1','fq0','fq+1','fq+2','fq+3','fq+4']
YH = ['fy-2','fy-1','fy0','fy+1','fy+2','fy+3','fy+4','fy+5']
HORN = {'fq+1':1,'fq+2':2,'fq+3':3,'fq+4':4,'fy+1':1,'fy+2':2,'fy+3':3,'fy+4':4,'fy+5':5}

# scale token -> multiplier that takes the raw cell to US$ MILLIONS. A BLANK scale on a money
# row means the cell is in whole units: AMZN's North-America operating income arrives as
# 9123000000 while its sibling segments arrive as 1717 / 16621 with scale 'M'. Reading the
# scale column is the only thing standing between that and a 1,000,000x error on one series.
SCALE = {'M': 1.0, 'B': 1000.0, 'K': 0.001, 'T': 1000000.0, '': 1e-6, None: 1e-6}

def num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s=str(v).strip()
    if not s or s.startswith('#N/A'): return None
    try: return float(s)
    except ValueError: return None

def canon(lbl):
    if lbl is None: return None, None
    s=str(lbl).strip()
    m=re.match(r'^(\d{4})\s+Q([1-4])\s*\((Rep|Fwd)\)',s)
    if m: return m.group(2)+'Q'+m.group(1)[2:], m.group(3)=='Rep'
    m=re.match(r'^(\d{4})\s+A\s*\((Rep|Fwd)\)',s)
    if m: return m.group(1), m.group(2)=='Rep'
    return None, None

def ordp(p): return int(p[2:])*4+int(p[0]) if 'Q' in p else int(p)

def segnorm(s):
    """'SEG0000227430 Segment' and 'SEG0000227430' are the same segment."""
    s = (s or '').strip()
    return s.split()[0].upper() if s else ''

def slot_index(r, idx):
    """This ROW's own declarations: (code, segment) -> (slot name, scale token).
    Built per row because slot numbering is a property of the export, not of the ticker."""
    out = {}
    for pre, tag in (('', 'metric'), ('_kpi', 'kpi')):
        for i in range(1, 51):
            cc = 'code%s%d' % (pre, i) if not pre else 'code_kpi%d' % i
            if cc not in idx or idx[cc] >= len(r): continue
            code = str(r[idx[cc]] or '').strip()
            if not code or code == '0.': continue
            sc_c = ('scale%d' % i) if not pre else ('scale_kpi%d' % i)
            sg_c = ('segment%d' % i) if not pre else ('segment_kpi%d' % i)
            scale = str(r[idx[sc_c]] or '').strip() if sc_c in idx and idx[sc_c] < len(r) else ''
            seg   = segnorm(r[idx[sg_c]]) if sg_c in idx and idx[sg_c] < len(r) else ''
            out[(code.upper(), seg)] = ('%s%d' % (tag, i), scale)
    return out

def spec_of(v):
    """Config value -> normalized spec dict. A bare string is a legacy slot name or a code."""
    if isinstance(v, str): return {'slot_or_code': v, 'unit': 'usdM', 'sign': 1}
    s = dict(v)
    s.setdefault('unit', 'usdM'); s.setdefault('sign', 1)
    return s

def resolve(spec, slots, idx):
    """-> (slot name, scale token) or None. Code+segment first; then a legacy slot name."""
    if 'code' in spec:
        seg = spec.get('segment', '')
        seg = segnorm(SEGS.get(seg, seg))
        return slots.get((spec['code'].upper(), seg))
    name = spec.get('slot_or_code', '')
    hit = slots.get((name.upper(), ''))
    if hit: return hit
    # legacy: the config named the column prefix itself (`kpi1`, or the pre-v2 `rev`)
    if any(k == name + '_' + h for k in idx for h in QH + YH): return (name, 'M')
    return None

def read_rows(rows, hdr, origin):
    idx={n:i for i,n in enumerate(hdr)}
    out=[]
    for r in rows:
        if not r or not r[0] or not str(r[0]).upper().startswith(TK+' '): continue
        asof=str(r[idx['data_as_of']])[:10]
        if '/' in asof:                                  # 8/4/2026 -> 2026-08-04
            mm,dd,yy = asof.split('/'); asof = '%s-%02d-%02d' % (yy, int(mm), int(dd))
        slots = slot_index(r, idx)
        fwd, lastQ, lastY = {}, None, None
        for h in QH+YH:
            if h not in idx: continue
            per,isrep = canon(r[idx[h]])
            if not per: continue
            if isrep:
                # Bloomberg's OWN reported column: never an estimate, and sometimes on
                # another basis than the company's measure. Used only to date the snapshot.
                if 'Q' in per: lastQ = per if (lastQ is None or ordp(per)>ordp(lastQ)) else lastQ
                else:          lastY = per if (lastY is None or ordp(per)>ordp(lastY)) else lastY
                continue
            for mkey, raw in list(MAP.get('q',{}).items()) + list(MAP.get('y',{}).items()):
                spec = spec_of(raw)
                hit = resolve(spec, slots, idx)
                if not hit: continue
                slot, scale = hit
                col = '%s_%s' % (slot, h)
                if col not in idx or idx[col] >= len(r): continue
                v = num(r[idx[col]])
                if v is None: continue
                unit = spec.get('unit','usdM')
                if unit == 'usdM': v *= SCALE.get(scale, 1e-6)
                v *= spec.get('sign', 1)
                fwd.setdefault(mkey,{})[per]=(v, HORN.get(h,9))
        out.append({'asof':asof,'fwd':fwd,'lastQ':lastQ,'lastY':lastY,'origin':origin,
                    'slots':slots})
    return out

rows_txt = list(csv.reader(open(TXT,newline='',encoding='utf-8-sig'),delimiter='\t'))
snaps = read_rows(rows_txt[1:], [str(v) for v in rows_txt[0]], 'txt')

wbk = openpyxl.load_workbook(WB, read_only=True, data_only=True)
sheet = next((s for s in WB_SHEETS if s in wbk.sheetnames), None)
if sheet is None:
    raise SystemExit('no consensus sheet in %s (looked for %s; found %s)' % (WB, WB_SHEETS, wbk.sheetnames))
it = wbk[sheet].iter_rows(values_only=True); wh=[str(v) for v in next(it)]
snaps += read_rows(list(it), wh, 'xlsm')

# dedupe by asof — archive wins on conflict (it is the immutable copy), but note collisions
by={}
for s in snaps:
    if s['asof'] in by:
        by[s['asof']]['origin'] += '+'+s['origin']
    else:
        by[s['asof']]=s
snaps=sorted(by.values(), key=lambda s:s['asof'])
print('%s: %d unique snapshots from %s + sheet %s' % (TK, len(snaps), os.path.basename(TXT), sheet))
only_txt  = [s['asof'] for s in snaps if s['origin']=='txt']
only_xlsm = [s['asof'] for s in snaps if s['origin']=='xlsm']
print('   archive-only: %s' % (', '.join(only_txt) or 'none'))
print('   sheet-only:   %s   <- lost if the .txt is not re-exported' % (', '.join(only_xlsm) or 'none'))

# ── resolution report: every configured metric, and whether the workbook can serve it ──
print('\nmetric resolution (against the newest snapshot):')
newest = snaps[-1] if snaps else None
missing = []
for view in ('q','y'):
    for mkey, raw in MAP.get(view,{}).items():
        spec = spec_of(raw); hit = resolve(spec, newest['slots'], {}) if newest else None
        n = sum(1 for s in snaps if s['fwd'].get(mkey))
        tag = ('%s scale=%r' % (hit[0], hit[1])) if hit else 'UNRESOLVED'
        print('   %-3s %-10s %-34s %-18s  %2d/%d snapshots' % (
            view, mkey, spec.get('code', spec.get('slot_or_code','')), tag, n, len(snaps)))
        if not hit: missing.append('%s.%s' % (view, mkey))
if missing: print('   !! unresolved: %s' % ', '.join(missing))

# ── the frozen-series check ───────────────────────────────────────────────────────────────────
# A genuine consensus is a fresh survey every quarter; it essentially never reprints a value to
# four decimals, let alone the same three values for five years. When a BQL cell dies, the export
# keeps emitting its last result and RELABELS it a year forward at each print — which reads as a
# perfectly smooth estimate series and is entirely fictional. AMZN's ANNUAL North-America revenue
# does exactly this: 473,099 / 517,634 / 561,728 at every snapshot since 2021-05, labelled FY21-23
# then FY22-24 ... then FY26-28. Its quarterly twin is fine, so this is per cell, not per ticker.
# Anything flagged here must be verified against Bloomberg and, until it is, dropped via `drop`.
print('\nfrozen-series check (a real consensus almost never reprints a value):')
frozen = []
for view in ('q','y'):
    for mkey in MAP.get(view, {}):
        cells = [repr(round(v[0], 4)) for s in snaps for p, v in s['fwd'].get(mkey, {}).items()
                 if ('Q' in p) == (view == 'q')]
        if not cells: continue
        uniq = len(set(cells))
        bad = uniq < len(cells) * 0.6
        if bad: frozen.append('%s.%s' % (view, mkey))
        print('   %s %s.%-9s %3d unique / %3d cells%s' % (
            '!!' if bad else '  ', view, mkey, uniq, len(cells), '   <-- FROZEN, do not ship' if bad else ''))
dropped = {v: set(MAP.get('drop', {}).get(v, [])) for v in ('q','y')}
if any(dropped.values()):
    print('   dropped by config: %s' % ', '.join(
        '%s.%s' % (v, k) for v in ('q','y') for k in sorted(dropped[v])))
still = [f for f in frozen if f.split('.')[1] not in dropped[f.split('.')[0]]]
if still: print('   !! FROZEN AND STILL BEING SHIPPED: %s — add them to "drop"' % ', '.join(still))

def jsnum(v):
    r=round(v,4)
    return str(int(r)) if r==int(r) else repr(r)

lines=[]
lines.append('  // ── GENERATED BLOCK — do not hand-edit. Rebuilt by scripts/consensus/emit_matrix.py')
lines.append('  // (see docs/RESULTS_CONVENTIONS.md §8). Source: the UNION of BBG_CONSENSUS.txt (the')
lines.append('  // exported archive) and Consensus_Portal.xlsm sheet %s (the live sheet, which' % sheet)
lines.append('  // can overwrite its most recent row), deduped by data_as_of. Only FORWARD horizons')
lines.append('  // are kept — the (Rep)-marked columns are Bloomberg\'s own reported figures, not')
lines.append('  // estimates, and can sit on a different basis than the company\'s own measure.')
lines.append('  estMatrix: {')
lines.append('    cons: {')
lines.append('      vintages: [')
for s in snaps:
    la = "{ q: %s, y: %s }" % (("'%s'"%s['lastQ']) if s['lastQ'] else 'null',
                               ("'%s'"%s['lastY']) if s['lastY'] else 'null')
    lines.append("        { id: '%s', label: '%s', lastActual: %s }," % (s['asof'], s['asof'], la))
lines.append('      ],')
for view, mp in (('q',MAP.get('q',{})), ('y',MAP.get('y',{}))):
    lines.append('      %s: {' % view)
    for mkey in mp:
        if mkey in dropped[view]: continue
        vfrom = (VALID.get(view,{}) or {}).get(mkey)
        cells=[]
        for s in snaps:
            d = s['fwd'].get(mkey, {})
            pers = [p for p in d if ('Q' in p) == (view=='q')]
            if vfrom: pers=[p for p in pers if ordp(p)>=ordp(vfrom)]
            if not pers: continue
            pers.sort(key=ordp)
            cells.append("        '%s': { %s }," % (s['asof'],
                ', '.join("'%s': %s" % (p, jsnum(d[p][0])) for p in pers)))
        if not cells: continue
        note = ("   // consensus only from %s — earlier BBG cells are off-basis"%vfrom) if vfrom else ''
        lines.append('      %s: {%s' % (mkey, note))
        lines += cells
        lines.append('      },')
    lines.append('      },')
lines.append('    }')
lines.append('  },')
NL = chr(10)
open(os.path.join(OUT, 'estmatrix_%s.js' % TK.lower()), 'w', encoding='utf-8', newline=NL).write(NL.join(lines) + NL)

# JSON sidecar so verify_preprint.py can replay the ENGINE's rule over exactly what was emitted,
# instead of re-deriving it from the workbook and testing a second implementation.
side = {'ticker': TK,
        'vintages': [{'id': s['asof'], 'label': s['asof'],
                      'lastActual': {'q': s['lastQ'], 'y': s['lastY']}} for s in snaps]}
for view in ('q','y'):
    side[view] = {}
    for mkey in MAP.get(view, {}):
        if mkey in dropped[view]: continue
        rows_ = {}
        for s in snaps:
            d = s['fwd'].get(mkey, {})
            cells = {p: round(v[0], 4) for p, v in d.items() if ('Q' in p) == (view=='q')}
            vfrom = (VALID.get(view,{}) or {}).get(mkey)
            if vfrom: cells = {p: v for p, v in cells.items() if ordp(p) >= ordp(vfrom)}
            if cells: rows_[s['asof']] = cells
        if rows_: side[view][mkey] = rows_
json.dump(side, open(os.path.join(OUT, 'estmatrix_%s.json' % TK.lower()), 'w', encoding='utf-8'), indent=1)
print('\nwrote out/estmatrix_%s.js  (%d lines)  + .json sidecar' % (TK.lower(), len(lines)))
